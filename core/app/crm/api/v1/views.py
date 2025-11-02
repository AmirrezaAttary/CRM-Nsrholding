from django.http import HttpResponse
from django.shortcuts import render
from django.shortcuts import redirect
from django.shortcuts import get_object_or_404

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.crm.models import CallReport, CargoAnnouncement, PurchaseProcess, SaleReport
from app.crm.api.v1.filters import CallReportFilter, SaleReportFilter
from app.crm.api.v1.serializer import (
    CallReportSerializer, 
    CargoAnnouncementSerializer, 
    PurchaseProcessSerializer, 
    SaleReportSerializer,
    SaleReportExportSerializer,
)
from app.crm.api.v1.paginations import CustomPagination
from app.crm.api.v1.permissions import CustomIsAuthenticated

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView



class CallReportViewSet(viewsets.ModelViewSet):
   
    queryset = CallReport.objects.all().order_by('-created_at')
    serializer_class = CallReportSerializer
    permission_classes = [CustomIsAuthenticated]

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = CallReportFilter  
    search_fields = ['name', 'number', 'province', 'city']
    ordering_fields = ['created_at', 'name']

    # Add custom pagination
    pagination_class = CustomPagination

    @action(detail=True, methods=['get'], url_path='go-to-purchase')
    def go_to_purchase(self, request, pk=None):
        call_report = self.get_object()
        purchase_process, created = PurchaseProcess.objects.get_or_create(call_report=call_report)
        purchase_url = f"/crm/api/v1/purchase-process/{purchase_process.id}/"
        return redirect(purchase_url)

class CargoAnnouncementViewSet(viewsets.ModelViewSet):
    queryset = CargoAnnouncement.objects.all()
    serializer_class = CargoAnnouncementSerializer
    permission_classes = [CustomIsAuthenticated]

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    
    filterset_fields = ['load_type', 'product_type__id', 'country_name__name']
    search_fields = ['product_type__name', 'country_name__name']
    ordering_fields = ['product_price', 'id']

    # Add custom pagination
    pagination_class = CustomPagination

class PurchaseProcessViewSet(viewsets.ModelViewSet):
    queryset = PurchaseProcess.objects.all().order_by('-created_at')
    serializer_class = PurchaseProcessSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['load_type', 'market_place_type']
    search_fields = ['call_report__name', 'call_report__number', 'buyer_name']
    ordering_fields = ['created_at', 'buyer_name', 'call_report__name']
    pagination_class = CustomPagination
    permission_classes = [CustomIsAuthenticated]

    @action(detail=True, methods=['get'], url_path='go-to-sale-report')
    def go_to_sale_report(self, request, pk=None):
        purchase_process = self.get_object()

        sale_type_value = getattr(purchase_process, 'load_type', None) or 'market_place'

        sale_report, created = SaleReport.objects.get_or_create(
            purchase_process=purchase_process,
            defaults={'sale_type': sale_type_value}
        )

        if not created and sale_report.sale_type != sale_type_value:
            sale_report.sale_type = sale_type_value
            sale_report.save(update_fields=['sale_type'])

        sale_report_url = f"/crm/api/v1/sale-reports/{sale_report.id}/"
        return redirect(sale_report_url)
    
    
#########################################################################

class SaleReportViewSet(viewsets.ModelViewSet):
    queryset = SaleReport.objects.all()
    serializer_class = SaleReportSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = SaleReportFilter
    pagination_class = CustomPagination
    permission_classes = [CustomIsAuthenticated]

##################################################

class SaleReportExportView(APIView):
    permission_classes = [CustomIsAuthenticated]
    def get(self, request, pk):
        sale_report = SaleReport.objects.get(pk=pk)
        serializer = SaleReportSerializer(sale_report, context={'request': request})
        data = serializer.data

        sale_type = data.get("sale_type")
        nested_data = None

        if sale_type == "market_place":
            nested_data = data.get("marketplace")
        elif sale_type == "market_outside":
            nested_data = data.get("marketoutside")
        elif sale_type == "quota":
            nested_data = data.get("quota")
        elif sale_type == "overhead":
            nested_data = data.get("overhead")

        FIELD_LABELS = {
            "sale_type" : "نوع خرید",
            "sale_date" : "تاریخ خرید",

            "market_place_product_name" : "نام کالا",
            "market_place_weight" : "وزن کالا",
            "market_place_market_price" : "قیمت بازارگاهی",
            "market_place_purchase_price" : "قیمت خرید",
            "market_place_selling_price" : "قیمت فروش",
            "market_place_profit" : "سود",
            "market_place_unofficial" : "غیررسمی",
            "market_place_total_amount" : "مبلغ کل",
            "market_place_deposit" : "واریزی",
            "market_place_account_remaining" : "مانده حساب",
            "market_place_buyer" : "خریدار",
            "market_place_seller" : "فروشنده",
            "market_place_supplier" : "عرضه کننده",
            "market_place_supply_status_name" : "وضعیت عرضه",
            "market_place_sales_expert_name" : "نام کارشناس فروش",
            "market_place_description" : "توضیحات",
            "market_place_weight_barname" : "وزن بارنامه",

            "market_outside_product_name" : "نام کالا" ,
            "market_outside_weight" : "وزن کالا" ,
            "market_outside_purchase_price" : "قیمت خرید" ,
            "market_outside_selling_price" : "قیمت فروش" ,
            "market_outside_profit" : "سود" ,
            "market_outside_total_amount" : "قیمت کل" ,
            "market_outside_deposit" : "واریزی" ,
            "market_outside_account_remaining" : "مانده حساب" ,
            "market_outside_buyer" : "خریداد" ,
            "market_outside_seller" : "فروشنده" ,
            "market_outside_supplier" : "عرضه کننده" ,
            "market_outside_supply_status_name" : "وضعیت عرضه" ,
            "market_outside_sales_expert_name" : "نام کارشناس فروش" ,
            "market_outside_description" : "توضیحات" ,
            "market_outside_weight_barname" : "وزن بارنامه" ,

            "quota_product_name" : "نام کالا" ,
            "quota_weight" : "وزن کالا" ,
            "quota_purchase_price" : "قیمت خرید" ,
            "quota_selling_price" : "قیمت فروش" ,
            "quota_profit" : "سود" ,
            "quota_total_amount" : "مبلغ کل" ,
            "quota_deposit" : "واریزی" ,
            "quota_account_remaining" : "مانده حساب" ,
            "quota_buyer" : "خریدار" ,
            "quota_seller" : "فروشنده" ,
            "quota_supplier" : "عرضه کننده" ,
            "quota_supply_status_name" : "وضعیت عرضه" ,
            "quota_sales_expert_name" : "نام کارشناس فروش" ,
            "quota_description" : "توضیحات" ,

            "overhead_address" : "آدرس" ,
            "overhead_number" : "شماره تماس" ,

        }

        EXCLUDE_FIELDS = ["id", "purchase_process", "sale_report", "supply_status", "export"]

        flat_data = {}
        for key, value in data.items():
            if key not in ["marketplace", "marketoutside", "quota", "overhead", "absolute_url"] + EXCLUDE_FIELDS:
                flat_data[key] = value

        if nested_data:
            for key, value in nested_data.items():
                prefixed_key = f"{sale_type}_{key}"
                if key not in EXCLUDE_FIELDS:
                    flat_data[prefixed_key] = value

        wb = Workbook()
        ws = wb.active
        ws.title = f"SaleReport_{pk}"

        
        headers = ["ردیف"] + [FIELD_LABELS.get(k, k) for k in flat_data.keys()]
        ws.append(headers)

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = green_fill

        SALE_TYPE_LABELS = {
            "market_place": "بازارگاهی",
            "market_outside": "خارج بازارگاهی",
            "quota": "سهمیه",
            "overhead": "روبار",
        }


        row_values = [1]  
        for k, v in flat_data.items():
            if k == "sale_type":
                row_values.append(SALE_TYPE_LABELS.get(v, v))
            else:
                row_values.append(str(v) if v is not None else "—")
        
        ws.append(row_values)

       
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=sale_report_{pk}.xlsx'
        wb.save(response)
        return response

class SaleReportBulkExportView(APIView):
    permission_classes = [CustomIsAuthenticated]
    def get(self, request):
        queryset = SaleReport.objects.all()
        sale_type_filter = request.GET.get("sale_type")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        if sale_type_filter:
            queryset = queryset.filter(sale_type=sale_type_filter)
        if start_date and end_date:
            queryset = queryset.filter(sale_date__range=[start_date, end_date])

        if not queryset.exists():
            return HttpResponse("هیچ گزارشی پیدا نشد.", status=404)

        FIELD_LABELS = {
            "sale_type" : "نوع خرید",
            "sale_date" : "تاریخ خرید",

            "market_place_product_name" : "نام کالا",
            "market_place_weight" : "وزن کالا",
            "market_place_market_price" : "قیمت بازارگاهی",
            "market_place_purchase_price" : "قیمت خرید",
            "market_place_selling_price" : "قیمت فروش",
            "market_place_profit" : "سود",
            "market_place_unofficial" : "غیررسمی",
            "market_place_total_amount" : "مبلغ کل",
            "market_place_deposit" : "واریزی",
            "market_place_account_remaining" : "مانده حساب",
            "market_place_buyer" : "خریدار",
            "market_place_seller" : "فروشنده",
            "market_place_supplier" : "عرضه کننده",
            "market_place_supply_status_name" : "وضعیت عرضه",
            "market_place_sales_expert_name" : "نام کارشناس فروش",
            "market_place_description" : "توضیحات",
            "market_place_weight_barname" : "وزن بارنامه",

            "market_outside_product_name" : "نام کالا" ,
            "market_outside_weight" : "وزن کالا" ,
            "market_outside_purchase_price" : "قیمت خرید" ,
            "market_outside_selling_price" : "قیمت فروش" ,
            "market_outside_profit" : "سود" ,
            "market_outside_total_amount" : "قیمت کل" ,
            "market_outside_deposit" : "واریزی" ,
            "market_outside_account_remaining" : "مانده حساب" ,
            "market_outside_buyer" : "خریداد" ,
            "market_outside_seller" : "فروشنده" ,
            "market_outside_supplier" : "عرضه کننده" ,
            "market_outside_supply_status_name" : "وضعیت عرضه" ,
            "market_outside_sales_expert_name" : "نام کارشناس فروش" ,
            "market_outside_description" : "توضیحات" ,
            "market_outside_weight_barname" : "وزن بارنامه" ,

            "quota_product_name" : "نام کالا" ,
            "quota_weight" : "وزن کالا" ,
            "quota_purchase_price" : "قیمت خرید" ,
            "quota_selling_price" : "قیمت فروش" ,
            "quota_profit" : "سود" ,
            "quota_total_amount" : "مبلغ کل" ,
            "quota_deposit" : "واریزی" ,
            "quota_account_remaining" : "مانده حساب" ,
            "quota_buyer" : "خریدار" ,
            "quota_seller" : "فروشنده" ,
            "quota_supplier" : "عرضه کننده" ,
            "quota_supply_status_name" : "وضعیت عرضه" ,
            "quota_sales_expert_name" : "نام کارشناس فروش" ,
            "quota_description" : "توضیحات" ,

            "overhead_address" : "آدرس" ,
            "overhead_number" : "شماره تماس" ,

        }

        SALE_TYPE_LABELS = {
            "market_place": "بازارگاهی",
            "market_outside": "خارج بازارگاهی",
            "quota": "سهمیه",
            "overhead": "روبار",
        }

        EXCLUDE_FIELDS = ["id", "purchase_process", "sale_report", "supply_status", "export"]

        wb = Workbook()
        ws = wb.active
        ws.title = "SaleReports"

        first_record = queryset.first()
        serializer = SaleReportSerializer(first_record, context={'request': request})
        sample_data = serializer.data

        flat_keys = []

        for record in queryset:
            serializer = SaleReportSerializer(record, context={'request': request})
            data = serializer.data

            sale_type = data.get("sale_type")
            nested_data = None
            nested_map = {
                "market_place": "marketplace",
                "market_outside": "marketoutside",
                "quota": "quota",
                "overhead": "overhead"
            }
            nested_data = data.get(nested_map.get(sale_type, ""), {})

            flat_data = {}
            for key, value in data.items():
                if key not in ["marketplace", "marketoutside", "quota", "overhead", "absolute_url"] + EXCLUDE_FIELDS:
                    flat_data[key] = value

            if nested_data:
                for key, value in nested_data.items():
                    prefixed_key = f"{sale_type}_{key}"
                    if key not in EXCLUDE_FIELDS:
                        flat_data[prefixed_key] = value

            for key in flat_data.keys():
                if key not in flat_keys:
                    flat_keys.append(key)

        headers = ["ردیف"] + [FIELD_LABELS.get(k, k) for k in flat_keys]
        ws.append(headers)

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.fill = green_fill

        row_number = 1
        for record in queryset:
            serializer = SaleReportSerializer(record, context={'request': request})
            data = serializer.data

            sale_type = data.get("sale_type")
            nested_data = data.get(nested_map.get(sale_type, ""), {})

            flat_data = {}
            for key, value in data.items():
                if key not in ["marketplace", "marketoutside", "quota", "overhead", "absolute_url"] + EXCLUDE_FIELDS:
                    flat_data[key] = value
            if nested_data:
                for key, value in nested_data.items():
                    prefixed_key = f"{sale_type}_{key}"
                    if key not in EXCLUDE_FIELDS:
                        flat_data[prefixed_key] = value

            row_values = [row_number]
            for k in flat_keys:
                v = flat_data.get(k, "—")
                if k == "sale_type":
                    v = SALE_TYPE_LABELS.get(v, v)
                row_values.append(str(v) if v is not None else "—")

            ws.append(row_values)
            row_number += 1

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=sale_reports.xlsx'
        wb.save(response)
        return response